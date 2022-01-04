import openpyxl


class ExcelInit():
    wb = None
    ws = None

    def __init__(self, filename, sheetname):
        self.filename = filename
        self.sheetname = sheetname

    def open_sheet(self):
        self.wb = openpyxl.load_workbook(self.filename)
        self.ws = self.wb[self.sheetname]

    def get_dimensions(self):
        return self.ws.max_row, self.ws.max_column

    def get_coordinates(self):
        for row in self.ws.iter_rows():
            for point in row:
                yield point.coordinate

    @staticmethod
    def compare_values(file1, file2, point):
        if file1.ws[point].value != file2.ws[point].value:
            print(f"Values in both sheet do not match at coordinate - {point}")
            print(f"value in {file1.ws} at {point} - {file1.ws[point].value}")
            print(f"value in {file2.ws} at {point} - {file2.ws[point].value}")
            print()

